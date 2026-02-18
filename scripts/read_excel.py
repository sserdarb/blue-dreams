import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\sserd\OneDrive\Belgeler\Antigravity\2026 Yönetim Raporu.xlsx', data_only=True)

# Focus on management-relevant sheets
key_sheets = [
    'Dash Board', 'SEZON RAPORU', 'HAFTALIK RAPOR',
    'Daily Report', 'Sales Report', 'Room Revenue',
    'CHANNEL', 'AGENCY', 'NATIONALITY', 'SEGMENTS',
    'Rate Plan Report', 'MICE TEKLIFLERI',
    'STOP SALE CHART', 'LAST 3 YEARS', 'Report Configuration'
]

# Also check all sheet names
all_names = wb.sheetnames

# Find sheets that exist
result = {}
for name in all_names:
    ws = wb[name]
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=False):
        cells = {}
        for c in row:
            if c.value is not None:
                cells[c.coordinate] = str(c.value)[:60]
        if cells:
            rows_data.append(cells)
    result[name] = {
        'rows': ws.max_row,
        'cols': ws.max_column,
        'header_rows': rows_data
    }

with open(r'C:\Users\sserd\OneDrive\Belgeler\Antigravity\blue-dreams-fix\scripts\excel_headers.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print("DONE")
